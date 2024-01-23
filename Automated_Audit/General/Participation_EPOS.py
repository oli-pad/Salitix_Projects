from collections import namedtuple
import pyodbc
import pandas as pd
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from collections import Counter

def Participation_EPOS(Promo_Schedule_df,Customer_Charges_df,EPOS_conn,EPOS_query,Client_Code,Retailer_Code):
    Exceptions_Report=[]
    for i in Promo_Schedule_df.index:
        row=Promo_Schedule_df.iloc[i]
        if row["Retailer_Product_Number"]=="TBC":break
        matched_cc_df,Invoice_list=Matching_CC(row,Customer_Charges_df)
        if matched_cc_df.empty==True:
            #Exceptions
            continue
        else:
            #print(EPOS_query.format(str(row["Retailer_Product_Number"]).replace(".0",""),row["Instore_Start_Date"],row["Instore_End_Date"]))
            EPOS_df=pd.read_sql(EPOS_query.format(str(row["Retailer_Product_Number"]).replace(".0",""),row["Instore_Start_Date"],row["Instore_End_Date"]),EPOS_conn)
            #rename columns if possible
            if EPOS_df.empty==True:
                EPOS_df=pd.read_sql(EPOS_query.format("0"+str(row["Retailer_Product_Number"]).replace(".0",""),row["Instore_Start_Date"],row["Instore_End_Date"]),EPOS_conn)
            try:
                EPOS_df.rename(columns = {'Sales_Value_TY':'Salitix_EPOS_Value'}, inplace = True)
                EPOS_df.rename(columns = {'Sales_Volume_TY':'Salitix_EPOS_Qty'}, inplace = True)
            except:None
            #print(EPOS_df)
            if row["Standard_Selling_Price"]==row["Promotional_Selling_Price"]:
                #print("Same")
                continue
            Participation_EPOS_df=Participated_EPOS(EPOS_df,row["Standard_Selling_Price"],row["Promotional_Selling_Price"])
            EPOS_total=Participation_EPOS_df["Promo_Units_Sold"].sum()
            CC_Qty_total=matched_cc_df["Quantity"].sum()
            difference = CC_Qty_total-EPOS_total
            if difference*float(row["Epos_Funding_Amount"])>50:
                Exceptions_Report.append(["'Not Reviewed'","'EPOS Discrepancy'","'"+str(row["Promo_period"]).replace("'","")+"'","'"+str(row["Retailer_Product_Number"])+"'","'"+str(round(difference*float(row["Epos_Funding_Amount"])))+"'","'"+str(row["Instore_Start_Date"])+"'","'"+str(row["Instore_End_Date"])+"'","'"+str(row["Instore_Start_Date"])+"'","'"+str(row["Instore_End_Date"])+"'","'"+str(row["Instore_Start_Date"])+"'","'"+str(row["Instore_End_Date"])+"'","NULL","'"+str(row["Retailer_promotion_number"])+"'",Invoice_list])
                #Claim_pack(str(row["Promo_period"])+" "+str(row["Retailer_Product_Number"])+" "+str(i)+" £"+str(round(difference*float(row["Epos_Funding_Amount"]))),Participation_EPOS_df,matched_cc_df,Promo_Schedule_df[Promo_Schedule_df["Retailer_Product_Number"]==row["Retailer_Product_Number"]],Client_Code,Retailer_Code)
            else:
                #Exceptions
                continue
    return Exceptions_Report

def Matching_CC(row,Customer_Charges_df):
    Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==float(row["Retailer_Product_Number"])]
    if Product_CC_df.empty==True:
        Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==row["Retailer_Product_Number"]]
    if Product_CC_df.empty==True:
        Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]=="0"+str(row["Retailer_Product_Number"])]
    if Product_CC_df.empty==True:
        Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==str(row["Retailer_Product_Number"]).replace("0","")]  
    #print(Product_CC_df)
    #print(Customer_Charges_df["Product_No"])
    Date_CC_df=Product_CC_df[Product_CC_df["End_Date"]>=row["Instore_Start_Date"]]
    Date_CC_df=Date_CC_df[Date_CC_df["Start_Date"]<=row["Instore_End_Date"]]
    if Date_CC_df.empty==True:
        return Date_CC_df,Invoice_list(Date_CC_df)
    Promo_Counts = Counter(Date_CC_df["Promotion_No"].tolist())
    Promo_Number = Promo_Counts.most_common(1)
    if len(Promo_Number)>1:
    
        comb_df=Date_CC_df[Date_CC_df["Promotion_No"]==Promo_Number[0][0]]
        for i in range(len(Promo_Number)):
            if i==0:continue
            comb_df=comb_df.append(Date_CC_df[Date_CC_df["Promotion_No"]==Promo_Number[0][i]])
        return comb_df
    else:
        Date_CC_df=Date_CC_df[Date_CC_df["Promotion_No"]==Promo_Number[0][0]]
        #Less than or equal to the dates on the customer charges
        return Date_CC_df,Invoice_list(Date_CC_df)

def Invoice_list(df):
    invoices=df["Invoice_No"].unique()
    string_list="'"
    invoices=sorted(invoices)
    for i in invoices:
       
        string_list+=str(i)+','
    string_list=replace_last(string_list,",","")
    string_list+="'"
    return string_list

def replace_last(string, find, replace):
    reversed = string[::-1]
    replaced = reversed.replace(find[::-1], replace[::-1], 1)
    return replaced[::-1]

def participation(Standard_RSP,Promo_RSP,Qty_Sold,Sales_Value):
    if Qty_Sold==0 or Sales_Value==0:
        return 1,0
    SV_Std_price=float(Standard_RSP)*float(Qty_Sold)
    SV_Pro_price=float(Promo_RSP)*float(Qty_Sold)
    Average_RSP=float(Sales_Value)/float(Qty_Sold)
    Full_Giveaway=SV_Std_price-SV_Pro_price
    Act_Giveaway=SV_Std_price-Sales_Value
    Promotional_partcipation=Act_Giveaway/Full_Giveaway
    if Promotional_partcipation>1:
        Promotional_partcipation=1
    elif Promotional_partcipation<0:
        Promotional_partcipation=0
    Promotional_Units=Promotional_partcipation*Qty_Sold
    return Promotional_partcipation,Promotional_Units

def Participated_EPOS(EPOS_df,Standard,Promo):
    Promotional_partcipation_list=[]
    Promotional_Units_list=[]
    Standard_RSP=Standard
    Promo_RSP=Promo
    for i in EPOS_df.index:
        Qty=EPOS_df["Salitix_EPOS_Qty"].iloc[i]
        Value=EPOS_df["Salitix_EPOS_Value"].iloc[i]
        Promotional_partcipation,Promotional_Units=participation(Standard_RSP,Promo_RSP,Qty,Value)
        Promotional_partcipation_list.append(Promotional_partcipation)
        Promotional_Units_list.append(Promotional_Units)
    EPOS_df["Promo_Participation"]=Promotional_partcipation_list
    EPOS_df["Promo_Units_Sold"]=Promotional_Units_list
    return EPOS_df


def Claim_pack(name,EPOS_df,matched_cc_df,Promos,Client_Code,Retailer_Code):
    wb = Workbook("Calculation")
    wb.create_sheet("CC")
    sheet = wb.active
    for row in dataframe_to_rows(matched_cc_df, index=False, header=True):
        sheet.append(row)
    wb.create_sheet("EPOS")
    ws = wb.worksheets[1]
    for row in dataframe_to_rows(EPOS_df, index=False, header=True):
        ws.append(row)
    wb.create_sheet("Promotions")
    ws = wb.worksheets[2]
    for row in dataframe_to_rows(Promos, index=False, header=True):
        ws.append(row)
    path_rename=r"C:\Users\Python\Desktop\General Analysis Project\Outcomes\{}\{}\Potential_Claims\EPOS_Discrepency"
    wb.save(os.path.join(path_rename.format(Retailer_Code,Client_Code),name+".xlsx"))
