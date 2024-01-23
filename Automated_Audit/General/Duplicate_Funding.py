from collections import namedtuple
import pyodbc
import pandas as pd
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from collections import Counter
import statistics

def Duplicate_Funding(Promo_Schedule_df,Customer_Charges_df,SI_conn,SI_query,Client_Code,Retailer_Code,EPOS_conn,EPOS_query):
    Exceptions_Report=[]
    Product_No_list=Promo_Schedule_df["Retailer_Product_Number"].unique()
    for i in Product_No_list:
        if i=="TBC":break
        Avg_Case_Cost_list=[]
        Mode_Case_Cost_list=[]
        prod_schedule_df=Promo_Schedule_df[Promo_Schedule_df["Retailer_Product_Number"]==i]
        for j in range(len(prod_schedule_df)):
            row=prod_schedule_df.iloc[j]
            matched_cc_df,Invoice_list=Matching_CC(row,Customer_Charges_df)
            if matched_cc_df.empty==True:
                None#Exceptions to be determined.
            else:
                EPOS_df=pd.read_sql(EPOS_query.format(row["Retailer_Product_Number"],row["Instore_Start_Date"],row["Instore_End_Date"]),EPOS_conn)
                #rename columns if possible
                if EPOS_df.empty==True:
                    EPOS_df=pd.read_sql(EPOS_query.format("0"+row["Retailer_Product_Number"],row["Instore_Start_Date"],row["Instore_End_Date"]),EPOS_conn)
                try:
                    EPOS_df.rename(columns = {'Sales_Value_TY':'Salitix_EPOS_Value'}, inplace = True)
                    EPOS_df.rename(columns = {'Sales_Volume_TY':'Salitix_EPOS_Qty'}, inplace = True)
                except:None
                Case_Cost_avg,Case_Cost_mode,SI_df=Case_Cost(row,SI_conn,SI_query)
                if SI_df.empty:
                    continue
                print(SI_df)
                Avg_Case_Cost_list.append(Case_Cost_avg)
                Mode_Case_Cost_list.append(Case_Cost_mode)
                if round(Case_Cost_mode,2)!=round(Case_Cost_avg,2):
                    difference=abs(Case_Cost_mode-Case_Cost_avg)
                    quantity=abs(SI_df["qty"].sum())
                    value=float(difference)*float(quantity)
                    if value>100000:value=50000
                    adj_Start=row["Instore_Start_Date"]-pd.DateOffset(days=15)
                    adj_End=row["Instore_End_Date"]+pd.DateOffset(days=15)
                    SI_df=pd.read_sql(SI_query.format("0"+str(row["Retailer_Product_Number"]),adj_Start,adj_End),SI_conn)
                    if SI_df.empty:
                        SI_df=pd.read_sql(SI_query.format(str(row["Retailer_Product_Number"]),adj_Start,adj_End),SI_conn)
                    #Claim_pack(str(row["Promo_period"])+" "+str(row["Retailer_Product_Number"])+" "+str(j),SI_df,matched_cc_df,prod_schedule_df,Client_Code,Retailer_Code,EPOS_df)
                    Exceptions_Report.append(["'Not Reviewed'","'Duplicate Funding'","'"+str(row["Promo_period"]).replace("'","")+"'","'"+str(row["Retailer_Product_Number"])+"'","'"+str(round(value))+"'","'"+str(row["Instore_Start_Date"])+"'","'"+str(row["Instore_End_Date"])+"'","'"+str(row["Instore_Start_Date"])+"'","'"+str(row["Instore_End_Date"])+"'","'"+str(row["Instore_Start_Date"])+"'","'"+str(row["Instore_End_Date"])+"'","NULL","'"+str(row["Retailer_promotion_number"])+"'",Invoice_list])
    return Exceptions_Report

def Matching_CC(row,Customer_Charges_df):
    Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==float(row["Retailer_Product_Number"])]
    if Product_CC_df.empty==True:
        Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==row["Retailer_Product_Number"]]
    if Product_CC_df.empty==True:
        Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]=="0"+str(row["Retailer_Product_Number"])]
    if Product_CC_df.empty==True:
        Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==str(row["Retailer_Product_Number"]).replace("0","")]  
    Date_CC_df=Product_CC_df[Product_CC_df["End_Date"]>=row["Instore_Start_Date"]]
    Date_CC_df=Date_CC_df[Date_CC_df["Start_Date"]<=row["Instore_End_Date"]]
    if Date_CC_df.empty==True:
        return Date_CC_df,Invoice_list(Date_CC_df)
    Promo_Counts = Counter(Date_CC_df["Promotion_No"].tolist())
    Promo_Number = Promo_Counts.most_common(1)
    if len(Promo_Number)>1:
        print(Promo_Number)
        comb_df=Date_CC_df[Date_CC_df["Promotion_No"]==Promo_Number[0][0]]
        for i in range(len(Promo_Number)):
            if i==0:continue
            comb_df=comb_df.append(Date_CC_df[Date_CC_df["Promotion_No"]==Promo_Number[0][i]])
        return comb_df
    else:
        Date_CC_df=Date_CC_df[Date_CC_df["Promotion_No"]==Promo_Number[0][0]]
        #Less than or equal to the dates on the customer charges
        return Date_CC_df,Invoice_list(Date_CC_df)

def Case_Cost(row,SI_conn,SI_query):
    SI_df=pd.read_sql(SI_query.format("0"+str(row["Retailer_Product_Number"]),row["Instore_Start_Date"],row["Instore_End_Date"]),SI_conn)
    if SI_df.empty:
        SI_df=pd.read_sql(SI_query.format(str(row["Retailer_Product_Number"]),row["Instore_Start_Date"],row["Instore_End_Date"]),SI_conn)
        if SI_df.empty:
            return False,False,SI_df
    Case_Costs=SI_df["unit_price"].tolist()
    Case_Cost_avg=sum(Case_Costs)/len(Case_Costs)
    try:
        Case_Cost_mode=statistics.mode(Case_Costs)
    except:
        Case_Cost_mode=1
    return Case_Cost_avg,Case_Cost_mode,SI_df

def Invoice_list(df):
    invoices=df["Invoice_No"].unique()
    string_list="'"
    invoices=sorted(invoices)
    for i in invoices:
        print(i)
        string_list+=str(i)+','
    string_list=replace_last(string_list,",","")
    string_list+="'"
    return string_list

def replace_last(string, find, replace):
    reversed = string[::-1]
    replaced = reversed.replace(find[::-1], replace[::-1], 1)
    return replaced[::-1]

def Claim_pack(name,SI_df,matched_cc_df,Promos,Client_Code,Retailer_Code,EPOS_df):
    wb = Workbook("Calculation")
    wb.create_sheet("Invoices")
    sheet = wb.active
    for row in dataframe_to_rows(matched_cc_df, index=False, header=True):
        sheet.append(row)
    wb.create_sheet("Sales Invoices")
    ws = wb.worksheets[1]
    for row in dataframe_to_rows(SI_df, index=False, header=True):
        ws.append(row)
    wb.create_sheet("Promotions")
    ws = wb.worksheets[2]
    for row in dataframe_to_rows(Promos, index=False, header=True):
        ws.append(row)
    wb.create_sheet("EPOS")
    ws = wb.worksheets[3]
    for row in dataframe_to_rows(EPOS_df, index=False, header=True):
        ws.append(row)
    path_rename=r"C:\Users\Python\Desktop\General Analysis Project\Outcomes\{}\{}\Potential_Claims\Duplicate_Funding"
    wb.save(os.path.join(path_rename.format(Retailer_Code,Client_Code),name+".xlsx"))
