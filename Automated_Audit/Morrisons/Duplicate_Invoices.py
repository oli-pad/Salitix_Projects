from collections import namedtuple
import pyodbc
import pandas as pd
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from collections import Counter
from itertools import combinations

def Duplicate_Invoices(Product_File_df,Customer_Charges_df,Client_Code):
    Products_list=Customer_Charges_df['Product_No'].unique()
    Date_list=Customer_Charges_df['Start_Date'].unique()
    for i in Products_list:
        prod_filter_df=Customer_Charges_df[Customer_Charges_df['Product_No']==i]
        for j in Date_list:
            prod_date_filter_df=prod_filter_df[prod_filter_df['Start_Date']==j]
            invoice_number_list=prod_date_filter_df['Invoice_No'].unique()
            if len(invoice_number_list)>1:
                Inv_dict={}
                for k in invoice_number_list:
                    Inv_df=prod_date_filter_df[prod_date_filter_df['Invoice_No']==k]
                    Inv_dict[k]=Inv_df['Net_Amount'].sum()
                comb=combinations(invoice_number_list,2)
                #comb=comb.tolist()
                for l in comb:
                    diff=float(Inv_dict[l[0]])-float(Inv_dict[l[1]])
                    if Inv_dict[l[0]]==0:continue
                    diff_perc=abs(diff)/float(Inv_dict[l[0]])
                    if diff_perc<0.10 and float(Inv_dict[l[0]])>500:
                        Claim_pack(str(l[0])+"_"+str(l[1]),prod_date_filter_df,Client_Code)
                    else:
                        None
                        #Exceptions list
            else:
                continue

def Claim_pack(name,matched_cc_df,Client_Code):
    wb = Workbook("Calculation")
    wb.create_sheet("CC")
    sheet = wb.active
    for row in dataframe_to_rows(matched_cc_df, index=False, header=True):
        sheet.append(row)
    path_rename=r"C:\Users\Python\Desktop\Morrisons Analysis Project\Outcomes\{}\Potential_Claims\Duplicate_Invoices"
    wb.save(os.path.join(path_rename.format(Client_Code),name+".xlsx"))
