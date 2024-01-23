from collections import namedtuple
import pyodbc
import pandas as pd
from datetime import datetime
import numpy as np
from pdf2image import convert_from_path
pop_path=r"C:\Users\Python\Desktop\poppler-0.68.0_x86\poppler-0.68.0\bin"
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime,timedelta

Calculation_Col=['PO_Number','Invoice_Number','Pricing_Date','Invoice_Date','Retailer_Product_Number','Vendor_Product_Number','Product_Description','Units_Per_Case','Qty','Unit_Price','Lines_Goods_Value','Current_Case_Cost','Difference','Claim']

conn = pyodbc.connect('DRIVER=SQL Server;SERVER=UKSALSQL02;DATABASE=Salitix_Master_Data;Trusted_Connection=Yes;UID=SALITIX\SQLSalitixAuditorUsers')
cursor = conn.cursor()

#For User Input
Client_Code=input("What's the Client Code?  >")
Retailer_Code=input("What's the Retailer Code?  >")
Date_period_1=input("Start date?  >")
Date_period_2=input("End date?  >")
#To store the name of the databse for the relevant client.
cursor.execute("SELECT Salitix_client_number,Salitix_client_name,Db FROM [Salitix_Master_Data].[dbo].[salitix_client_numbers] WHERE Salitix_client_number='"+Client_Code+"';")
User_List=cursor.fetchall()
Client_Db=User_List[0][2]

#To connect to sales invoices in the database
SI_conn = pyodbc.connect('DRIVER=SQL Server;SERVER=UKSALSQL02;DATABASE='+Client_Db+';Trusted_Connection=Yes;UID=SALITIX\SQLSalitixAuditorUsers')
SI_cursor = SI_conn.cursor()

#To store the sales invoices in the order in which is required for claims
query="SELECT order_number,invoice_number,pricing_date,billing_date,Retailer_Primary_Product_Number,Vendor_Primary_Product_Number,product_description,units_per_case,qty,unit_price,line_goods_value FROM ["+Client_Db+"].[dbo].[vw_SAL_Sales_Invoice_Detail] WHERE salitix_customer_number='"+Retailer_Code+"' AND pricing_date BETWEEN '"+Date_period_1+"' AND '"+Date_period_2+"';"
Sales_Invoices_df=pd.read_sql(query,SI_conn)

#Dataframe for product name Desc and End Date.
Price_df=pd.DataFrame(columns=["Product_Number","Product_Description","Start_Date","End_Date","Range"])


def Summary_tab(Sales_Invoices_df,Price_df):
    Product_list=Sales_Invoices_df["product_description"].unique()
    for i in Product_list:
        Index_match_df=Sales_Invoices_df[Sales_Invoices_df["product_description"]==i]
        pricing_date_list=Index_match_df["pricing_date"].tolist()
        Start_Date=Date_period_1
        End_Date=max(pricing_date_list)
        Start_Date=datetime(day=int(Start_Date[3:5]),month=int(Start_Date[0:2]),year=int(Start_Date[6:10]))
        timedelta=End_Date-Start_Date
        Price_df.loc[len(Price_df.index)]=[Index_match_df["Retailer_Primary_Product_Number"].iloc[0],i,Start_Date,End_Date,int(timedelta.days)]
    Create_Exceptions(Price_df,Sales_Invoices_df)

def Create_Exceptions(df,df2):
    wb = Workbook("Summary")
    wb.create_sheet("Summary")
    sheet = wb.active
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)
    wb.create_sheet("Sales_Invoices")
    ws = wb.worksheets[1]
    for row in dataframe_to_rows(df2, index=False, header=True):
        ws.append(row)
    wb.save(os.path.join(r"C:\Users\Python\Desktop\projects\Sales_Invoice_Analysis\docs",r"Exceptions.xlsx"))

Summary_tab(Sales_Invoices_df,Price_df)
