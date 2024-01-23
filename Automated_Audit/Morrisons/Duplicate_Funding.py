from collections import namedtuple
import pyodbc
import pandas as pd
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from collections import Counter
import statistics

def Duplicate_Funding(Promo_Schedule_df,Product_File_df,Customer_Charges_df,SI_conn,SI_query,Client_Code):
    Product_No_list=Promo_Schedule_df["Item_number"].unique()
    for i in Product_No_list:
        Avg_Case_Cost_list=[]
        Mode_Case_Cost_list=[]
        prod_schedule_df=Promo_Schedule_df[Promo_Schedule_df["Item_number"]==i]
        for j in range(len(prod_schedule_df)):
            row=prod_schedule_df.iloc[j]
            min_number=Product_File_df[Product_File_df["Invoice_Description"]==int(row["Item_number"])]
            if min_number.empty==True:
                continue#Exceptions to be determined.
            else:
                matched_cc_df=Matching_CC(row,min_number["Product_No"],Customer_Charges_df)
                if matched_cc_df.empty==True:
                    None#Exceptions to be determined.
                else:
                    Case_Cost_avg,Case_Cost_mode,SI_df=Case_Cost(row,min_number["Product_No"],SI_conn,SI_query)
                    if SI_df.empty:
                        continue
                    Avg_Case_Cost_list.append(Case_Cost_avg)
                    Mode_Case_Cost_list.append(Case_Cost_mode)
                    if round(Case_Cost_mode,2)!=round(Case_Cost_avg,2):
                        print(Case_Cost_avg)
                        print(Case_Cost_mode)
                        Claim_pack(str(j)+"_"+str(min_number["Product_No"].values[0]),SI_df,matched_cc_df,prod_schedule_df,Client_Code)
        #for j in prod_schedule_df.index:
        #    if Case_Cost_list[j]!=Assumed_case_cost:
        #        row=prod_schedule_df.iloc[j]
        #        min_number=Product_File_df[Product_File_df["Invoice_Description"]==int(row["Item_number"])]
        #        if min_number.empty==True:
        #            continue#Exceptions to be determined.
        #        else:
        #            matched_cc_df=Matching_CC(row,min_number["Product_No"],Customer_Charges_df)
        #            if matched_cc_df.empty==True:
        #                None#Exceptions to be determined.
        #            else:
        #                Claim_pack(str(min_number["Product_No"])+"_"+str(row["Start_Date"]),SI_df,matched_cc_df,Promos)


def Matching_CC(row,min_number,Customer_Charges_df):
    Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==int(min_number)]
    Date_CC_df=Product_CC_df[Product_CC_df["End_Date"]>=row["Start_Date"]]
    Date_CC_df=Date_CC_df[Date_CC_df["Start_Date"]<=row["End_Date"]]
    if Date_CC_df.empty==True:
        return Date_CC_df
    elif "Supermarket" not in Date_CC_df["Store_Format"].tolist():
        return pd.DataFrame()
    Promo_Counts = Counter(Date_CC_df["Promotion_No"].tolist())
    Promo_Number = Promo_Counts.most_common(1)
    Product_CC_df=Product_CC_df[Product_CC_df["Promotion_No"]==Promo_Number[0][0]]
    #Less than or equal to the dates on the customer charges
    return Product_CC_df

def Case_Cost(row,min_number,SI_conn,SI_query):
    adj_Start=row["Start_Date"]-pd.DateOffset(days=15)
    adj_End=row["End_Date"]+pd.DateOffset(days=15)
    SI_df=pd.read_sql(SI_query.format("0"+str(min_number.values[0]),row["Start_Date"],row["End_Date"]),SI_conn)
    if SI_df.empty:
        return False,False,SI_df
    Case_Costs=SI_df["unit_price"].tolist()
    Case_Cost_avg=sum(Case_Costs)/len(Case_Costs)
    try:
        Case_Cost_mode=statistics.mode(Case_Costs)
    except:
        Case_Cost_mode=1
    SI_df=pd.read_sql(SI_query.format("0"+str(min_number.values[0]),adj_Start,adj_End),SI_conn)
    return Case_Cost_avg,Case_Cost_mode,SI_df

def Claim_pack(name,SI_df,matched_cc_df,Promos,Client_Code):
    wb = Workbook("Calculation")
    wb.create_sheet("Invoices")
    sheet = wb.active
    for row in dataframe_to_rows(matched_cc_df, index=False, header=True):
        sheet.append(row)
    wb.create_sheet("Sales Invoices")
    ws = wb.worksheets[1]
    for row in dataframe_to_rows(SI_df, index=False, header=True):
        ws.append(row)
    wb.create_sheet("Promotions")
    ws = wb.worksheets[2]
    for row in dataframe_to_rows(Promos, index=False, header=True):
        ws.append(row)
    path_rename=r"C:\Users\Python\Desktop\Morrisons Analysis Project\Outcomes\{}\Potential_Claims\Duplicate_Funding"
    wb.save(os.path.join(path_rename.format(Client_Code),name+".xlsx"))
