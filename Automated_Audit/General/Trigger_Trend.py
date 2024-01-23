from collections import namedtuple
import pyodbc
import pandas as pd
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from collections import Counter

def Trigger_Trend(Promo_Schedule_df,Customer_Charges_df,EPOS_conn,EPOS_query,Client_Code,Retailer_Code):
    Exceptions_Report=[]
    Product_No_list=Promo_Schedule_df["Retailer_Product_Number"].unique()
    for i in Product_No_list:
        prod_schedule_df=Promo_Schedule_df[Promo_Schedule_df["Retailer_Product_Number"]==i]
        trigger_list=prod_schedule_df["Epos_Funding_Amount"].unique()
        Start_Date_list=prod_schedule_df["Instore_End"].tolist()
        End_Date_list=prod_schedule_df["Instore_Start"].tolist()
        if len(trigger_list)>1:
            EPOS_df=pd.read_sql(EPOS_query.format(str(i),min(Start_Date_list),max(End_Date_list)),EPOS_conn)
            #rename columns if possible
            if EPOS_df.empty==True:
                EPOS_df=pd.read_sql(EPOS_query.format("0"+str(i),min(Start_Date_list),max(End_Date_list)),EPOS_conn)
            try:
                EPOS_df.rename(columns = {'Sales_Value_TY':'Salitix_EPOS_Value'}, inplace = True)
                EPOS_df.rename(columns = {'Sales_Volume_TY':'Salitix_EPOS_Qty'}, inplace = True)
            except:None
            matched_cc_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==float(i)]
            Exceptions_Report.append(["Not Reviewed","Trigger Trend","All",str(i),None,min(Start_Date_list),max(End_Date_list),None,None,None,None,None,None])
            #Claim_pack(str(i).replace(".0",""),matched_cc_df,prod_schedule_df,Client_Code,Retailer_Code,EPOS_df)
    return Exceptions_Report


def Claim_pack(name,matched_cc_df,Promos,Client_Code,Retailer_Code,EPOS_df):
    wb = Workbook("Calculation")
    wb.create_sheet("CC")
    sheet = wb.active
    for row in dataframe_to_rows(matched_cc_df, index=False, header=True):
        sheet.append(row)
    wb.create_sheet("Promotions")
    ws = wb.worksheets[1]
    for row in dataframe_to_rows(Promos, index=False, header=True):
        ws.append(row)
    wb.create_sheet("EPOS")
    ws = wb.worksheets[2]
    for row in dataframe_to_rows(EPOS_df, index=False, header=True):
        ws.append(row)
    path_rename=r"C:\Users\Python\Desktop\General Analysis Project\Outcomes\{}\{}\Potential_Claims\Trigger_Trend"
    wb.save(os.path.join(path_rename.format(Retailer_Code,Client_Code),name+".xlsx"))
