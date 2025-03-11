import openpyxl
from openpyxl.utils import FORMULAE
import re

input_file = "C:\\Users\\python\\Desktop\\projects\\Split_excel_file_python\\Pricing_review_test.xlsx"  # Specify your input Excel file
output_prefix = "output_sheet"  # Prefix for output Excel files
sheet1_name = "01.11.21-31.10.22 Sales Invoice"  # Name of the sheet 1 to separate line by line
sheet2_name = "Summary"  # Name of the sheet 2 to keep the same

def separate_excel(input_file, output_prefix, sheet1_name, sheet2_name):
    # Read the Excel file
    workbook = openpyxl.load_workbook(input_file, data_only=True)
    sheet1 = workbook[sheet1_name]
    sheet2 = workbook[sheet2_name]

    # Iterate over each row in sheet 1 and create separate files
    for index, row in enumerate(sheet1.iter_rows(min_row=2, values_only=True), start=1):
        output_file = f"{output_prefix}_{index}.xlsx"
        new_workbook = openpyxl.Workbook()

        # Add header row from Sheet1 to the new workbook
        new_sheet1 = new_workbook.active
        new_sheet1.title = sheet1_name
        headers = [cell.value for cell in sheet1[1]]
        for col_idx, header in enumerate(headers, start=1):
            new_sheet1.cell(row=1, column=col_idx, value=header)

        # Add data row from Sheet1 to the new workbook
        for col_idx, value in enumerate(row, start=1):
            new_sheet1.cell(row=2, column=col_idx, value=value)

        # Copy Sheet2 to the new workbook
        new_sheet2 = new_workbook.create_sheet(title=sheet2_name)
        for row_idx in range(1, sheet2.max_row + 1):
            for col_idx in range(1, sheet2.max_column + 1):
                new_sheet2.cell(row=row_idx, column=col_idx, value=sheet2.cell(row=row_idx, column=col_idx).value)

        # Update cell D40 in the copied Sheet2 to reference M2 in Sheet1
        new_sheet2['D40'].value = f'={sheet1_name}!M2'

        # Save the new file
        new_workbook.save(output_file)

# Call the function
separate_excel(input_file, output_prefix, sheet1_name, sheet2_name)

#import pandas as pd

#input_file = "C:\\Users\\python\\Desktop\\projects\\Split_excel_file_python\\Pricing_review_test.xlsx"  # Specify your input Excel file
#output_prefix = "output_sheet"  # Prefix for output Excel files
#rows_per_sheet = 1  # Number of rows per sheet

#def separate_excel(input_file, output_prefix, rows_per_sheet):
    # Read the Excel file
    #df = pd.read_excel(input_file)
    
    # Determine the number of sheets needed
    #num_sheets = len(df) // rows_per_sheet
    #if len(df) % rows_per_sheet != 0:
        #num_sheets += 1
    
    # Split the DataFrame into chunks and save as separate Excel files
    #for i in range(num_sheets):
        #start_idx = i * rows_per_sheet
        #end_idx = min((i + 1) * rows_per_sheet, len(df))
        #sheet_df = df.iloc[start_idx:end_idx]
        #output_file = f"{output_prefix}_{i + 1}.xlsx"
        #sheet_df.to_excel(output_file, index=False)
        #Print(output_file)

# Call the function
#separate_excel(input_file, output_prefix, rows_per_sheet)

