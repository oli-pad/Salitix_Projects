import requests
import pandas as pd
import re
import time
from datetime import date
from datetime import timedelta
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
from pathlib import Path
import shutil
from collections import namedtuple
from sys import argv
from openpyxl import load_workbook
from openpyxl import Workbook
import csv

driver = webdriver.Chrome()
days=["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
months={'01':"January","02":"February","03":"March","04":"April","05":"May","06":"June","07":"July","08":"August","09":"September","10":"October","11":"November","12":"Decemeber"}
short_months={"January":"Jan","February":"Feb","March":"Mar","April":"Apr","May":"May","June":"Jun","July":"Jul","August":"Aug","September":"Sep","October":"Oct","November":"Nov","Decemeber":"Dec"}
folder_name={"BACARDI":"Bacardi","AB_INBEV":"Ab_Inbev","AG_BARR":"AG_Barr","COTY":"Coty","FBC":"Fbc","FINSBURY_FOODS":"Finsbury_Foods","FBC":"Fbc",
"HEINEKEN":"Heineken","KETTLE_FOODS":"Kettle_foods","KINNERTON":"Kinnerton","KP_SNACKS":"KP_Snacks","MAXXIUM":"Maxxium","PLADIS":"Pladis",
"PREMIER_FOODS":"Premier_foods","TILDA":"Tilda","WEETABIX":"Weetabix","YOUNGS":"Youngs","PRINCES":"Princes","LOREAL":"Loreal"}
###Log in page####       
Tesco_toolkit=r'https://partnerstoolkit.tesco.com/sign-in/?url=%2Ftoolkit%2F'
EPOS_URL=r'https://partnerstoolkit.tesco.com/partner/reports/'

Line = namedtuple('Line',"Client_Name File_Name Status")
report=[]
lines=[]

script,Client_Name,User_Email,User_Password,number_of_days,start_num=argv

def get_download_data(Client_Name,User_Email,User_Password):
    time.sleep(5)
    download_data(Client_Name,User_Email,User_Password)
    time.sleep(5)
    logoff()

def download_data(client_name,username,password):
    selecting_dates,selecting_downloads=dates_for_download(int(number_of_days))
    login(Tesco_toolkit,username,password)
    time.sleep(5)
    print("Files to be download: ")
    print(selecting_downloads)
    print("Downloading files for {}".format(client_name))
    download_files(selecting_downloads)
    #CSV_to_Excels()
    print("Renaming files for {}".format(client_name))
    rename_relocate(client_name)

def dates_for_download(n):
    date_list=[]
    download_list=[]
    today = date.today()-timedelta(int(start_num))
    for i in range(n):
        Dates=today-timedelta(1+i)
        date_one = Dates.strftime("%d/%m/%Y")
        day=Dates.weekday()
        if date_one[0] == "0":
            date_list.append([date_one[1],months[date_one[3:5]],date_one[6:],days[day]])
        else:
            date_list.append([date_one[:2],months[date_one[3:5]],date_one[6:],days[day]])
    for j in date_list:
        download_list.append('Tesco-Sales_and_stock-TPNB_Sales_x_store-'+str(j[2])+short_months[j[1]]+str(j[0])+".xlsx")
    return date_list,download_list

# OO changes 03.05.2024
# Delete all files in the Downloads folder that start with "Tesco-Sales_and_stock-TPNB_Sales_x_store-"
def delete_files():
    for file in os.listdir(os.path.join(os.path.expanduser('~'), 'Downloads')):
        if file.startswith("Tesco-Sales_and_stock-TPNB_Sales_x_store-"):
            os.remove(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),file))

def login(url,username,password):
   delete_files()
   driver.get(url)
   time.sleep(1)
   WebDriverWait(driver,150).until(EC.presence_of_element_located((By.ID,"partner-signin-btn")))
   driver.find_element(By.ID,"partner-signin-btn").click()
   time.sleep(1)
   WebDriverWait(driver,150).until(EC.presence_of_element_located((By.ID,"email-address")))
   driver.find_element(By.ID,"email-address").send_keys(username)
   driver.find_element(By.ID,"Password").send_keys(password)
   time.sleep(1)
   driver.find_element(By.XPATH,"//button[@class='styled__StyledTextButton-sc-8hxn3m-0 iaukuD styled__StyledButton-NtoXt iBoFMG ddsweb-button ddsweb-button--text-button']").click()

def download_files(selecting_downloads):
    time.sleep(20)
    driver.get("https://partnerstoolkit.tesco.com/partner/reports/stored")
    WebDriverWait(driver,60).until(EC.presence_of_element_located((By.ID,"downloadedReports")))
    driver.find_element(By.ID,'downloadedReports').click()
    time.sleep(60) # reduce time from 200 to 60
    # Can't find the page source
    soup = BeautifulSoup(driver.page_source,"html.parser")
    links=soup.find_all("a") # Links is HTML code that starts with "a" which are the "a class" links.
    selecting_downloads=[i.replace(".xlsx",".csv") for i in selecting_downloads] # Changes .xlsx to .csv so code can find it in HTML code
    #DownloadCount = 0 # JR added int variable to fix bug. If the portal loads without loading generated file links for download, the "links" variable won't hold any values and the code will skip the while loop.
    for i in links: # Loops through every "a class" HTML code (download links)
        if i.text in selecting_downloads: # if the text of the HTML code is in the selecting_downloads list, it will enter the while loop
            while True:
                try:
                    time.sleep(30)
                    driver.find_element(By.ID,i.get('id')).click()
                    print("Attempting to download file: " + i.text)
                    time.sleep(70)
                    driver.find_element(By.ID,i.get('id')) # Waits for the portal to still be on the same page. If not, it will raise an exception.
                    print ("Portal still on the same page") # TEMP FOR DEBUG
                    #DownloadCount += 1
                    break # iterates to next i element
                except Exception as e:
                    print("Error in downloading file " + i.text)
                    time.sleep(10)
                    driver.refresh() # Refresh page added to fix the issue of the page staying on the downloaded reports tab but reports not showing
                    WebDriverWait(driver,30).until(EC.presence_of_element_located((By.ID,"downloadedReports")))
                    driver.find_element(By.ID,'downloadedReports').click()
                    continue # stays on this i element until the file is downloaded
        #if DownloadCount == 0:
        #    download_files(selecting_downloads) # If no files are downloaded, it will call the function again to try to download the files again

def CSV_to_Excels():
    time.sleep(40)
    selecting_dates,selecting_downloads=dates_for_download(int(number_of_days))
    selecting_downloads=[i.replace(".xlsx",".csv") for i in selecting_downloads]
    for i in selecting_downloads:
        try:
            df=pd.read_csv(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),i),sep=None)
            df.to_excel(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),i.replace(".csv",".xlsx")),sheet_name='Store Level LDI',index=False)
            os.remove(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),i))
        except:None

def rename_relocate(client_name):
    time.sleep(30)
    selecting_dates,selecting_downloads=dates_for_download(int(number_of_days))
    rename_path="Z:\{}\EPOS Store\Tesco\Daily"
    file_list=dir_contents(rename_path.format(folder_name[client_name]))
    # if client_name == "MAXXIUM":
    #     for i in selecting_downloads:
    #         wb=load_workbook(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),i))
    #         ws=wb['Store Level LDI'].values
    #         columns = next(ws)[0:]
    #         df = pd.DataFrame(ws,columns=columns)
    #         to_drop=[]
    #         for j in df.index:
    #             if df.loc[j].at["Supplier"]==36308 or df.loc[j].at["Supplier"]==59365:
    #                 to_drop.append(j)
    #         df=df.drop(to_drop)
    #         df.to_csv(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),i),sheet_name='Store Level LDI',index=False)
    # if client_name == "KETTLE_FOODS":
    #     for i in selecting_downloads:
    #         wb=load_workbook(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),i))
    #         ws=wb['Store Level LDI'].values
    #         columns = next(ws)[0:]
    #         df = pd.DataFrame(ws,columns=columns)
    #         to_drop=[]
    #         for j in df.index:
    #             if df.loc[j].at["Supplier"]==61814 or df.loc[j].at["Supplier"]==59365:
    #                 to_drop.append(j)
    #         df=df.drop(to_drop)
    #         df.to_csv(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),i),sheet_name='Store Level LDI',index=False)
    # else:None
    selecting_downloads=[i.replace(".xlsx",".csv") for i in selecting_downloads]
    for i in selecting_downloads:
        #try:
        print("i= ",i)
        name=i.replace(".csv"," {}.csv".format(folder_name[client_name])) # Puts client name at the end of file name
        print("name= ",name)
        time.sleep(10)
        if name not in file_list:
            os.rename(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),i),os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),name)) # Replace the file name with the new name (i with name)
            shutil.move(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),name),os.path.join(rename_path.format(folder_name[client_name]),name)) # Move the file to the new folder Z:\{}\EPOS Store\Tesco\Daily
        else:
            os.rename(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),i),os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),name)) # Replace the file name with the new name (i with name)
            os.replace(os.path.join(os.path.join(os.path.expanduser('~'), 'Downloads'),name),os.path.join(rename_path.format(folder_name[client_name]),name)) # Replace the file in the new folder Z:\{}\EPOS Store\Tesco\Daily
        #except:None

def dir_contents(path):
    file_list=[]
    for root, directories, files in os.walk(path):
        for filename in files:
            file_list.append(files)
    return file_list

def logoff():
    time.sleep(30)
    driver.find_element(By.ID,"profile").click()
    time.sleep(2)
    driver.find_element(By.ID,"signout").click()
    time.sleep(2)

get_download_data(Client_Name,User_Email,User_Password)
driver.quit()
