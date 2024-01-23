from collections import namedtuple
import pyodbc
import pandas as pd
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from collections import Counter

Line=namedtuple('Line',"Start_Date Item_number Descritpion Comment")
Exceptions=[]

def Incorrect_Trigger(Promo_Schedule_df,Product_File_df,Customer_Charges_df,EPOS_conn,EPOS_query,Client_Code):
    for i in Promo_Schedule_df.index:
        row=Promo_Schedule_df.iloc[i]
        min_number=Product_File_df[Product_File_df["Invoice_Description"]==int(row["Item_number"])]
        if min_number.empty==True:
            Exceptions.append(Line(row["Start_Date"],row["Item_number"],row["Product_Description"],"No Product Link"))
            continue
        else:
            matched_cc_df=Matching_CC(row,min_number["Product_No"],Customer_Charges_df)
            if matched_cc_df.empty==True:
                Exceptions.append(Line(row["Start_Date"],row["Item_number"],row["Product_Description"],"No Customer Charges"))
            else:
                unit_funding=matched_cc_df["Unit_Price"].tolist()
                for j in unit_funding:
                    print(j)
                    print(row["Funding_per_Unit"])
                    if float(j) == float(row["Funding_per_Unit"]):
                        Exceptions.append(Line(row["Start_Date"],row["Item_number"],row["Product_Description"],"Correct unit price"))
                        continue
                    else:
                        if float(j)>=float(row["Funding_per_Unit"]):
                            Claim_pack(str(min_number["Product_No"].values[0])+" "+str(i),matched_cc_df,Promo_Schedule_df[Promo_Schedule_df["Item_number"]==row["Item_number"]],Client_Code)
                            break
    df=pd.DataFrame(Exceptions)
    df.to_csv(r'C:\Users\Python\Desktop\Morrisons Analysis Project\Outcomes\CL014\Exceptions\Exceptions_Incorrect_Trigger.csv',index=False)

def Matching_CC(row,min_number,Customer_Charges_df):
    Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==int(min_number)]
    Date_CC_df=Product_CC_df[Product_CC_df["End_Date"]>=row["Start_Date"]]
    Date_CC_df=Date_CC_df[Date_CC_df["Start_Date"]<=row["End_Date"]]
    if Date_CC_df.empty==True:
        return Date_CC_df
    elif "Supermarket" not in Date_CC_df["Store_Format"].tolist():
        return pd.DataFrame()
    Promo_Counts = Counter(Date_CC_df["Promotion_No"].tolist())
    Promo_Number = Promo_Counts.most_common(1)
    Product_CC_df=Product_CC_df[Product_CC_df["Promotion_No"]==Promo_Number[0][0]]
    #Less than or equal to the dates on the customer charges
    return Product_CC_df

def Claim_pack(name,matched_cc_df,Promos,Client_Code):
    wb = Workbook("Calculation")
    wb.create_sheet("CC")
    sheet = wb.active
    for row in dataframe_to_rows(matched_cc_df, index=False, header=True):
        sheet.append(row)
    wb.create_sheet("Promotions")
    ws = wb.worksheets[1]
    for row in dataframe_to_rows(Promos, index=False, header=True):
        ws.append(row)
    path_rename=r"C:\Users\Python\Desktop\Morrisons Analysis Project\Outcomes\{}\Potential_Claims\Incorrect_Trigger"
    wb.save(os.path.join(path_rename.format(Client_Code),name+".xlsx"))
