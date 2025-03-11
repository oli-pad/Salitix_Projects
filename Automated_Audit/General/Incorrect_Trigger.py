from collections import namedtuple
import pyodbc
import pandas as pd
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from collections import Counter

def Incorrect_Trigger(Promo_Schedule_df,Customer_Charges_df,EPOS_conn,EPOS_query,Client_Code,Retailer_Code):
    Exceptions_Report=[]
    for i in Promo_Schedule_df.index:
        row=Promo_Schedule_df.iloc[i]
        if row["Retailer_Product_Number"]=="TBC":break
        matched_cc_df,invoice_list=Matching_CC(row,Customer_Charges_df)
        date_promo_schedule=Promo_Schedule_df[Promo_Schedule_df["Instore_Start_Date"]==row["Instore_Start_Date"]]
        prod_date_promo_schedule=date_promo_schedule[date_promo_schedule["Retailer_Product_Number"]==row["Retailer_Product_Number"]]
        prod_date_unit_list=prod_date_promo_schedule["Epos_Funding_Amount"].tolist()
        if matched_cc_df.empty==True:
            #Exceptions
            continue
        else:
            EPOS_df=pd.read_sql(EPOS_query.format(row["Retailer_Product_Number"],row["Instore_Start_Date"],row["Instore_End_Date"]),EPOS_conn)
            #rename columns if possible
            if EPOS_df.empty==True:
                EPOS_df=pd.read_sql(EPOS_query.format("0"+row["Retailer_Product_Number"],row["Instore_Start_Date"],row["Instore_End_Date"]),EPOS_conn)
            try:
                EPOS_df.rename(columns = {'Sales_Value_TY':'Salitix_EPOS_Value'}, inplace = True)
                EPOS_df.rename(columns = {'Sales_Volume_TY':'Salitix_EPOS_Qty'}, inplace = True)
            except:None
            unit_funding=matched_cc_df["Unit_Price"].tolist()
            for j in unit_funding:
                if round(float(j),2) <= round(float(row["Epos_Funding_Amount"]),2) or j==0 or row["Epos_Funding_Amount"]==0:
                    #Exceptions
                    continue
                else:
                    if j not in prod_date_unit_list and round(float(j),2) not in prod_date_unit_list:
                        CC_total=matched_cc_df["Quantity"].sum()
                        claim_total=(float(row["Epos_Funding_Amount"])-float(j))*float(CC_total)
                        Exceptions_Report.append(["'Not Reviewed'","'Incorrect Unit Funding'","'"+str(row["Promo_period"])+"'","'"+str(row["Retailer_Product_Number"])+"'","'"+str(claim_total)+"'","'"+str(row["Instore_Start_Date"])+"'","'"+str(row["Instore_End_Date"])+"'","'"+str(row["Instore_Start_Date"])+"'","'"+str(row["Instore_End_Date"])+"'","'"+str(row["Instore_Start_Date"])+"'","'"+str(row["Instore_End_Date"])+"'","NULL","'"+str(row["Retailer_promotion_number"])+"'",invoice_list])
                        #Claim_pack(str(row["Promo_period"])+" "+str(row["Retailer_Product_Number"])+" "+str(i),matched_cc_df,Promo_Schedule_df[Promo_Schedule_df["Retailer_Product_Number"]==row["Retailer_Product_Number"]],Client_Code,Retailer_Code,EPOS_df)
                        break
    return Exceptions_Report

def Matching_CC(row,Customer_Charges_df):
    Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==float(row["Retailer_Product_Number"])]
    if Product_CC_df.empty==True:
        Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==row["Retailer_Product_Number"]]
    if Product_CC_df.empty==True:
        Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]=="0"+str(row["Retailer_Product_Number"])]
    if Product_CC_df.empty==True:
        Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==str(row["Retailer_Product_Number"]).replace("0","")]   
    Date_CC_df=Product_CC_df[Product_CC_df["End_Date"]>=row["Instore_Start_Date"]]
    Date_CC_df=Date_CC_df[Date_CC_df["Start_Date"]<=row["Instore_End_Date"]]
    if Date_CC_df.empty==True:
        return Date_CC_df,Invoice_list(Date_CC_df)
    Promo_Counts = Counter(Date_CC_df["Promotion_No"].tolist())
    Promo_Number = Promo_Counts.most_common(1)
    if len(Promo_Number)>1:
        comb_df=Date_CC_df[Date_CC_df["Promotion_No"]==Promo_Number[0][0]]
        for i in range(len(Promo_Number)):
            if i==0:continue
            comb_df=comb_df.append(Date_CC_df[Date_CC_df["Promotion_No"]==Promo_Number[0][i]])
        return comb_df,Invoice_list(comb_df)
    else:
        Date_CC_df=Date_CC_df[Date_CC_df["Promotion_No"]==Promo_Number[0][0]]
        #Less than or equal to the dates on the customer charges
        return Date_CC_df,Invoice_list(Date_CC_df)

def Invoice_list(df):
    invoices=df["Invoice_No"].unique()
    string_list="'"
    invoices=sorted(invoices)
    for i in invoices:
        print(i)
        string_list+=str(i)+','
    string_list=replace_last(string_list,",","")
    string_list+="'"
    return string_list

def replace_last(string, find, replace):
    reversed = string[::-1]
    replaced = reversed.replace(find[::-1], replace[::-1], 1)
    return replaced[::-1]

def Claim_pack(name,matched_cc_df,Promos,Client_Code,Retailer_Code,EPOS_df):
    wb = Workbook("Calculation")
    wb.create_sheet("CC")
    sheet = wb.active
    for row in dataframe_to_rows(matched_cc_df, index=False, header=True):
        sheet.append(row)
    wb.create_sheet("Promotions")
    ws = wb.worksheets[1]
    for row in dataframe_to_rows(Promos, index=False, header=True):
        ws.append(row)
    wb.create_sheet("EPOS")
    ws = wb.worksheets[2]
    for row in dataframe_to_rows(EPOS_df, index=False, header=True):
        ws.append(row)
    path_rename=r"C:\Users\Python\Desktop\General Analysis Project\Outcomes\{}\{}\Potential_Claims\Incorrect_Trigger"
    wb.save(os.path.join(path_rename.format(Retailer_Code,Client_Code),name+".xlsx"))
