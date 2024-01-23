from collections import namedtuple
import pyodbc
import pandas as pd
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from collections import Counter
from itertools import combinations

from Incorrect_Trigger import Invoice_list,replace_last


def Duplicate_Invoices(Customer_Charges_df,Client_Code,Retailer_Code,EPOS_conn,EPOS_query,Promo_Schedule_df):
    Exceptions_Report=[]
    Products_list=Customer_Charges_df['Product_No'].unique()
    Net_Amounts=Customer_Charges_df['Net_Amount'].unique()
    # for i in Net_Amounts:
    #    Net_filter_df=Customer_Charges_df[Customer_Charges_df['Net_Amount']==i]
    #    l=Net_filter_df['Invoice_No'].unique()
    #    Net_filter_df['Start_Date']=pd.to_datetime(Net_filter_df['Start_Date'])
    #    date=Net_filter_df['Start_Date'].tolist()
    #    print(date)
    #    if len(l)>1 and float(i)>800:
    #        invoices=l
    #        Exceptions_Report.append(["'Not Reviewed'","'Invoice Duplicates'","NULL","NULL","'"+str(i)+"'","'"+str(date[0])[:10]+"'","'"+str(date[0])[:10]+"'","'"+str(date[0])[:10]+"'","'"+str(date[0])[:10]+"'","'"+str(date[0])[:10]+"'","'"+str(date[0])[:10]+"'","NULL","NULL",str(Invoice_list(invoices))])
           #Claim_pack(str(l[0]).replace("/","_")+"_"+str(l[1]).replace("/","_"),Net_filter_df,Client_Code,Retailer_Code,Net_filter_df,Net_filter_df)
    for i in Products_list:
        prod_filter_df=Customer_Charges_df[Customer_Charges_df['Product_No']==i]
        prod_filter_df=prod_filter_df.sort_values(by="Start_Date")
        prod_filter_df['Start_Date'] = pd.to_datetime(prod_filter_df['Start_Date'])
        prod_filter_df['End_Date'] = pd.to_datetime(prod_filter_df['End_Date'])
        inv_list=prod_filter_df['Invoice_No'].tolist()
        Date_list=prod_filter_df['Start_Date'].tolist()
        End_Date_list=prod_filter_df['End_Date'].tolist()
        count=-1
        for j in Date_list:
            count+=1
            prod_date_filter_df=prod_filter_df[prod_filter_df['Start_Date']==j]
            invoice_number_list=prod_date_filter_df['Invoice_No'].unique()
            try:
                EPOS_df=pd.read_sql(EPOS_query.format(str(i).replace(".0",""),j,End_Date_list[count]),EPOS_conn)
                #rename columns if possible
                if EPOS_df.empty==True:
                    EPOS_df=pd.read_sql(EPOS_query.format("0"+str(i).replace(".0",""),j,End_Date_list[count]),EPOS_conn)
            except:continue
            try:
                EPOS_df.rename(columns = {'Sales_Value_TY':'Salitix_EPOS_Value'}, inplace = True)
                EPOS_df.rename(columns = {'Sales_Volume_TY':'Salitix_EPOS_Qty'}, inplace = True)
            except:None
            if len(invoice_number_list)>1:
                Inv_dict={}
                for k in invoice_number_list:
                    Inv_df=prod_date_filter_df[prod_date_filter_df['Invoice_No']==k]
                    Inv_dict[k]=Inv_df['Net_Amount'].sum()
                comb=combinations(invoice_number_list,2)
                #comb=comb.tolist()
                for l in comb:
                    diff=float(Inv_dict[l[0]])-float(Inv_dict[l[1]])
                    if Inv_dict[l[0]]==0:continue
                    diff_perc=abs(diff)/float(Inv_dict[l[0]])
                    if diff_perc<0.10 and float(Inv_dict[l[0]])>500:
                        EPOS_total=EPOS_df["Salitix_EPOS_Qty"].sum()
                        CC_Qty_total=prod_date_filter_df["Quantity"].sum()
                        difference = CC_Qty_total-EPOS_total
                        if difference/CC_Qty_total>0.1:
                            invoices=list(Inv_dict)
                            Exceptions_Report.append(["'Not Reviewed'","'Invoice Duplicates'","NULL","NULL","'"+str(Inv_dict[l[0]])+"'","'"+str(End_Date_list[count])+"'","'"+str(End_Date_list[count])+"'","'"+str(End_Date_list[count])+"'","'"+str(End_Date_list[count])+"'","'"+str(End_Date_list[count])+"'","'"+str(End_Date_list[count])+"'","NULL","NULL",str(Invoice_list(invoices))])
                            #Claim_pack(str(l[0]).replace("/","_")+"_"+str(l[1]).replace("/","_"),prod_date_filter_df,Client_Code,Retailer_Code,EPOS_df,Promo_Schedule_df)  
                    else:
                        None
                        #Exceptions list
            else:
                continue
        count=-1
        for j in End_Date_list:
            count+=1
            prod_date_filter_df=prod_filter_df[prod_filter_df['End_Date']==j]
            invoice_number_list=prod_date_filter_df['Invoice_No'].unique()
            try:
                EPOS_df=pd.read_sql(EPOS_query.format(str(i).replace(".0",""),Date_list[count],j),EPOS_conn)
                #rename columns if possible
                if EPOS_df.empty==True:
                    EPOS_df=pd.read_sql(EPOS_query.format("0"+str(i).replace(".0",""),Date_list[count],j),EPOS_conn)
            except:continue
            try:
                EPOS_df.rename(columns = {'Sales_Value_TY':'Salitix_EPOS_Value'}, inplace = True)
                EPOS_df.rename(columns = {'Sales_Volume_TY':'Salitix_EPOS_Qty'}, inplace = True)
            except:None
            if len(invoice_number_list)>1:
                Inv_dict={}
                for k in invoice_number_list:
                    Inv_df=prod_date_filter_df[prod_date_filter_df['Invoice_No']==k]
                    Inv_dict[k]=Inv_df['Net_Amount'].sum()
                comb=combinations(invoice_number_list,2)
                #comb=comb.tolist()
                for l in comb:
                    diff=float(Inv_dict[l[0]])-float(Inv_dict[l[1]])
                    if Inv_dict[l[0]]==0:continue
                    diff_perc=abs(diff)/float(Inv_dict[l[0]])
                    if diff_perc<0.10 and float(Inv_dict[l[0]])>500:
                        EPOS_total=EPOS_df["Salitix_EPOS_Qty"].sum()
                        CC_Qty_total=prod_date_filter_df["Quantity"].sum()
                        difference = CC_Qty_total-EPOS_total
                        if difference/CC_Qty_total>0.1:
                            value_exceptions=[Exceptions_Report[x][4] for x in range(len(Exceptions_Report))]
                            if "'"+str(i)+"'" in value_exceptions:continue
                            invoices=list(Inv_dict)
                            if len(invoices)>10:
                                invoices=invoices[0:10]
                            #Claim_pack(str(l[0]).replace("/","_")+"_"+str(l[1]).replace("/","_"),prod_date_filter_df,Client_Code,Retailer_Code,EPOS_df,Promo_Schedule_df)
                            Exceptions_Report.append(["'Not Reviewed'","'Invoice Duplicates'","NULL","NULL","'"+str(Inv_dict[l[0]])+"'","'"+str(End_Date_list[count])+"'","'"+str(End_Date_list[count])+"'","'"+str(End_Date_list[count])+"'","'"+str(End_Date_list[count])+"'","'"+str(End_Date_list[count])+"'","'"+str(End_Date_list[count])+"'","NULL","NULL",str(Invoice_list(invoices))])
                    else:
                        None
                        #Exceptions list
            else:
                continue
        # if Retailer_Code!="MOR01" or Retailer_Code!="SAI01":
        #     for x in range(len(Date_list)):
        #         if x==0:continue
        #         if Date_list[x]<=End_Date_list[x-1] and Date_list[x]!=Date_list[x-1]:
        #             try:
        #                 EPOS_df=pd.read_sql(EPOS_query.format(str(i).replace(".0",""),Date_list[x],End_Date_list[x]),EPOS_conn)
        #                 #rename columns if possible
        #                 if EPOS_df.empty==True:
        #                     EPOS_df=pd.read_sql(EPOS_query.format("0"+str(i).replace(".0",""),Date_list[x],End_Date_list[x]),EPOS_conn)
        #             except:continue
        #             sum_claims=EPOS_df["Salitix_EPOS_Qty"].sum()/30
        #             invoices = [inv_list[x].replace("/","_"),inv_list[x-1].replace("/","_")]
        #             #Claim_pack(str(inv_list[x]).replace("/","_")+"_"+str(inv_list[x-1]).replace("/","_")+" Overlap",prod_filter_df,Client_Code,Retailer_Code,EPOS_df,Promo_Schedule_df)
        #             Exceptions_Report.append(["'Not Reviewed'","'Overlapping Promotions'","NULL","NULL","'"+str(sum_claims)+"'","'"+str(min(Date_list))+"'","'"+str(max(End_Date_list))+"'","'"+str(min(Date_list))+"'","'"+str(max(End_Date_list))+"'","'"+str(min(Date_list))+"'","'"+str(max(End_Date_list))+"'","NULL","NULL",str(Invoice_list(invoices))])
    return Exceptions_Report


def Invoice_list(invoices):
    string_list="'"
    invoices=sorted(invoices)
    for i in invoices:
        print(i)
        
        string_list+=str(i)+','
    string_list=replace_last(string_list,",","")
    string_list+="'"
    return string_list

def replace_last(string, find, replace):
    reversed = string[::-1]
    replaced = reversed.replace(find[::-1], replace[::-1], 1)
    return replaced[::-1]

def Claim_pack(name,matched_cc_df,Client_Code,Retailer_Code,EPOS_df,Promo_Schedule_df):
    wb = Workbook("Calculation")
    wb.create_sheet("CC")
    sheet = wb.active
    for row in dataframe_to_rows(matched_cc_df, index=False, header=True):
        sheet.append(row)
    wb.create_sheet("EPOS")
    ws = wb.worksheets[1]
    for row in dataframe_to_rows(EPOS_df, index=False, header=True):
        ws.append(row)
    wb.create_sheet("Promotions")
    ws = wb.worksheets[2]
    for row in dataframe_to_rows(Promo_Schedule_df, index=False, header=True):
        ws.append(row)
    path_rename=r"C:\Users\Python\Desktop\General Analysis Project\Outcomes\{}\{}\Potential_Claims\Duplicate_Invoices"
    wb.save(os.path.join(path_rename.format(Retailer_Code,Client_Code),name+".xlsx"))
