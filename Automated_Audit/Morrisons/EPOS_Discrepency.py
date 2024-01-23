from collections import namedtuple
import pyodbc
import pandas as pd
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from collections import Counter

Line=namedtuple('Line',"Start_Date Item_number Descritpion Comment")
Exceptions=[]

def EPOS_Discrepency(Promo_Schedule_df,Product_File_df,Customer_Charges_df,EPOS_conn,EPOS_query,Client_Code):
    for i in Promo_Schedule_df.index:
        row=Promo_Schedule_df.iloc[i]
        min_number=Product_File_df[Product_File_df["Invoice_Description"]==int(row["Item_number"])]
        print(min_number)
        if min_number.empty==True:
            Exceptions.append(Line(row["Start_Date"],row["Item_number"],row["Product_Description"],"No Product Link"))
            continue
        else:
            matched_cc_df=Matching_CC(row,min_number["Product_No"],Customer_Charges_df)
            print(matched_cc_df)
            if matched_cc_df.empty==True:
                Exceptions.append(Line(row["Start_Date"],row["Item_number"],row["Product_Description"],"No Customer Charges"))
            else:
                EPOS_df=pd.read_sql(EPOS_query.format(min_number["Product_No"].values[0],row["Start_Date"],row["End_Date"]),EPOS_conn)
                EPOS_total=EPOS_df["Salitix_EPOS_Qty"].sum()
                CC_Qty_total=matched_cc_df["Quantity"].sum()
                if float(EPOS_total)<float(CC_Qty_total):
                    Claim_pack(str(min_number["Product_No"].values[0])+" "+str(i),EPOS_df,matched_cc_df,Promo_Schedule_df[Promo_Schedule_df["Item_number"]==row["Item_number"]],Client_Code)
                else:
                    Exceptions.append(Line(row["Start_Date"],row["Item_number"],row["Product_Description"],"EPOS too high"))
    df=pd.DataFrame(Exceptions)
    df.to_csv(r'C:\Users\Python\Desktop\Morrisons Analysis Project\Outcomes\CL014\Exceptions\Exceptions_EPOS_Discrepency.csv',index=False)

def Matching_CC(row,min_number,Customer_Charges_df):
    print(min_number)
    Product_CC_df=Customer_Charges_df[Customer_Charges_df["Product_No"]==int(min_number)]
    Date_CC_df=Product_CC_df[Product_CC_df["End_Date"]>=row["Start_Date"]]
    Date_CC_df=Date_CC_df[Date_CC_df["Start_Date"]<=row["End_Date"]]
    Date_CC_df=Date_CC_df[Date_CC_df["Store_Format"]=="Supermarket"]
    if Date_CC_df.empty==True:
        return Date_CC_df
    Promo_Counts = Counter(Date_CC_df["Promotion_No"].tolist())
    Promo_Number = Promo_Counts.most_common(1)
    Date_CC_df=Date_CC_df[Date_CC_df["Promotion_No"]==Promo_Number[0][0]]
    #Less than or equal to the dates on the customer charges
    return Date_CC_df

def Claim_pack(name,EPOS_df,matched_cc_df,Promos,Client_Code):
    wb = Workbook("Calculation")
    wb.create_sheet("CC")
    sheet = wb.active
    for row in dataframe_to_rows(matched_cc_df, index=False, header=True):
        sheet.append(row)
    wb.create_sheet("EPOS")
    ws = wb.worksheets[1]
    for row in dataframe_to_rows(EPOS_df, index=False, header=True):
        ws.append(row)
    wb.create_sheet("Promotions")
    ws = wb.worksheets[2]
    for row in dataframe_to_rows(Promos, index=False, header=True):
        ws.append(row)
    path_rename=r"C:\Users\Python\Desktop\Morrisons Analysis Project\Outcomes\{}\Potential_Claims\EPOS_Discrepency"
    wb.save(os.path.join(path_rename.format(Client_Code),name+".xlsx"))
