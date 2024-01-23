from collections import namedtuple
import pyodbc
import pandas as pd
from datetime import datetime
import numpy as np
from pdf2image import convert_from_path
pop_path=r"C:\Users\Python\Desktop\poppler-0.68.0_x86\poppler-0.68.0\bin"
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def remove_zero(string):
    while string[0] == '0':
        string = string[1:]
    return string


Calculation_Col=['PO_Number','Invoice_Number','Pricing_Date','Invoice_Date','Retailer_Product_Number','Vendor_Product_Number','Product_Description','Units_Per_Case','Qty','Unit_Price','Lines_Goods_Value','Current_Case_Cost','Difference','Claim']

conn = pyodbc.connect('DRIVER=SQL Server;SERVER=UKSALSQL02;DATABASE=Salitix_Master_Data;Trusted_Connection=Yes;UID=SALITIX\SQLSalitixAuditorUsers')
cursor = conn.cursor()

#For User Input
Client_Code=input("What's the Client Code?  >")
Retailer_Code=input("What's the Retailer Code?  >")
Date_period_1=input("Start date?  >")
Date_period_2=input("End date?  >")
#To store the name of the databse for the relevant client.
cursor.execute("SELECT Salitix_client_number,Salitix_client_name,Db FROM [Salitix_Master_Data].[dbo].[salitix_client_numbers] WHERE Salitix_client_number='"+Client_Code+"';")
User_List=cursor.fetchall()
Client_Db=User_List[0][2]

#To connect to sales invoices in the database
SI_conn = pyodbc.connect('DRIVER=SQL Server;SERVER=UKSALSQL02;DATABASE=Sandbox;Trusted_Connection=Yes;UID=SALITIX\SQLSalitixAuditorUsers')
SI_cursor = SI_conn.cursor()

#To store the sales invoices in the order in which is required for claims
query="SELECT order_number,invoice_number,pricing_date,billing_date,Retailer_Primary_Product_Number,Vendor_Primary_Product_Number,product_description,units_per_case,qty,unit_price,line_goods_value,product_number FROM ["+Client_Db+"].[dbo].[vw_SAL_Sales_Invoice_Detail] WHERE salitix_customer_number='"+Retailer_Code+"'AND line_goods_value>0 AND pricing_date BETWEEN '"+Date_period_1+"' AND '"+Date_period_2+"' ORDER BY pricing_date;"
Sales_Invoices_df=pd.read_sql(query,SI_conn)
Sales_Invoices_df["product_number"]=[remove_zero(i) for i in Sales_Invoices_df["product_number"]]
print(Sales_Invoices_df["product_number"])
#To connect to price file in the database
#PF_conn = pyodbc.connect('DRIVER=SQL Server;SERVER=UKSALSQL02;DATABASE=Salitix_Pricing_Data_Processed;Trusted_Connection=Yes;UID=SALITIX\SQLSalitixAuditorUsers')
#PF_cursor = PF_conn.cursor()
PF_conn = pyodbc.connect('DRIVER=SQL Server;SERVER=UKSALSQL02;DATABASE=Salitix_Pricing_Data_Formatted;Trusted_Connection=Yes;UID=SALITIX\SQLSalitixAuditorUsers')
PF_cursor = PF_conn.cursor()

#To store the price file
if Retailer_Code=="SAI01":
    query_PF="SELECT * FROM [Salitix_Pricing_Data_Formatted].[dbo].[Pricing_Sainsburys_Ftd] WHERE salitix_customer_number='"+Retailer_Code+"' OR salitix_client_number='"+Client_Code+"' ORDER BY Effective_Date, SourceFile DESC ;"
    PF_df=pd.read_sql(query_PF,PF_conn)
elif Retailer_Code=="TES01":
    query_PF="SELECT * FROM [Salitix_Pricing_Data_Formatted].[dbo].[Pricing_Tesco_Ftd] WHERE salitix_customer_number='"+Retailer_Code+"' AND salitix_client_number='"+Client_Code+"' ORDER BY Cost_Effective_Date, SourceFile DESC ;"
    PF_df=pd.read_sql(query_PF,PF_conn)
    PF_df.columns = PF_df.columns.str.replace('TPNB', 'Product_No')
    PF_df.columns = PF_df.columns.str.replace('Cost_Effective_Date', 'Effective_Date')
    PF_df.columns = PF_df.columns.str.replace('Final_Invoice_Cost', 'Std_case_cost')
    PF_df.columns = PF_df.columns.str.replace('Case_Size', 'Std_case_size')
    print(PF_df)
elif Retailer_Code=="COO01":
    query_PF="SELECT * FROM [Salitix_Pricing_Data_Staging].[dbo].[Pricing_Coop_stg] WHERE Salitix_Customer_Number='"+Retailer_Code+"' AND Salitix_Client_Number='"+Client_Code+"' ORDER BY Effective_Date;"
    PF_df=pd.read_sql(query_PF,PF_conn)
    PF_df.columns = PF_df.columns.str.replace('Invoice_Match_Price', 'Std_case_cost')
    print(PF_df)
###

##format for folder structure where pricing is considered
folder="W:\Audit\Bacardi\Documentation\Sainsbury_Salesforce\Price\{}"

#Functions for analysis
#The following will search the price file for what the correct price on each individual invoice.
def Insert_Correct_Case_Cost(SI_df,PF_df):
    Current_Case_Cost_col,Difference_col,Claim_col,SourceFile_col=[],[],[],[]
    for i in SI_df.index:
        print(i," out of ",len(SI_df), ", ",i*100/len(SI_df),"%")
        #Selects and individual sales invoice line
        # Prod_No=SI_df.at[i,'Retailer_Primary_Product_Number']
        Prod_No=SI_df.at[i,'product_number']
           
        if Prod_No:Prod_No=Prod_No.replace(" ","")
        locate=PF_df.loc[PF_df['Product_No']==Prod_No]
        
        Current_Case_Cost,Difference,Claim,SourceFile="No product link","No product link",0,"No product link"
        #Stores all the price infomation for the product number on given sales invoice line
        for j in range(len(locate)):
            Date_1=locate["Effective_Date"].iloc[j]
            Date_2=SI_df["pricing_date"].iloc[i]
            Date_1,Date_2=datetime.strptime(str(Date_1)[:10], '%Y-%m-%d'),datetime.strptime(str(Date_2)[:10], '%Y-%m-%d')
            if Date_1<=Date_2:
                Current_Case_Cost=locate["Std_case_cost"].iloc[j]
                #Current_Case_Size=locate["Std_case_size"].iloc[j]
                #if Current_Case_Size != SI_df.at[i,"units_per_case"]:
                #    Current_Case_Cost="Case Size is different"
                try:
                    if SI_df.at[i,"unit_price"]==0:
                        Difference=0
                        Claim=0
                    else:
                        Difference=-float(SI_df.at[i,"unit_price"])+float(Current_Case_Cost)
                        Claim=float(Difference)*float(SI_df.at[i,"qty"])
                except:
                    Difference,claim="Error",0
                SourceFile=(locate["SourceFile"].iloc[j])
                #When price date isn't less than or equal cost effect date then you'll have the correct price line.
                continue
            else:
                #Will now have the pricing line associated with the SI line
                break
        Current_Case_Cost_col.append(Current_Case_Cost)
        Difference_col.append(Difference)
        Claim_col.append(Claim)
        SourceFile_col.append(SourceFile)
    SI_df=SI_df.assign(Current_Case_Cost=Current_Case_Cost_col,Difference=Difference_col,Claim=Claim_col,SourceFile=SourceFile_col)
    return SI_df

#Look for where there is a positive claim amount, seperate this by product and then put product and total into a table
def Claim_Identifier(SI_A_df):
    df1=SI_A_df[SI_A_df["Claim"]>0]
    Claim_df=df1.sort_values(['product_number','SourceFile','pricing_date'],ascending = [True, True, True])
    table=pd.pivot_table(Claim_df,values="Claim",index=['product_number','SourceFile'],aggfunc=np.sum)
    print(table)
    #table.to_excel(os.path.join("W:\Audit\Bacardi\Documentation\Pricing_Claims\Summary",Client_Code+"_"+Retailer_Code+"_"+Date_period_1.replace("/",".")+"-"+Date_period_2.replace("/",".")+".xlsx"))
    seperate_claims = Claim_df.groupby(['product_number','SourceFile'])
    seperate_claims_series=[seperate_claims.get_group(x) for x in seperate_claims.groups]
    for i in seperate_claims_series:
        SourceFile_pdf=i["SourceFile"].iloc[0]
        i.pop("SourceFile")
        if Retailer_Code=="SAI01":
            i.columns=Calculation_Col
            i.pop("PO_Number")
            filename=i["Retailer_Product_Number"].iloc[0]+str(i["Pricing_Date"].iloc[0])[:10]+'.xlsx'
            Evidence_list=Evidence_func(SourceFile_pdf)
            insert_image(filename,Evidence_list,i)
        elif Retailer_Code=="TES01":
            i.columns=Calculation_Col
            filename=i["Retailer_Product_Number"].iloc[0]+str(i["Pricing_Date"].iloc[0])[:10]+'.xlsx'
            Prod_No=i["Retailer_Product_Number"].iloc[0]
            PF_df_ev=pd.read_sql(query_PF,PF_conn)
            Evidence=PF_df_ev.loc[PF_df_ev["TPNB"]==Prod_No.replace(" ","")]
            wb = Workbook("Calculation")
            wb.create_sheet("Calculation")
            sheet = wb.active
            for row in dataframe_to_rows(i, index=False, header=True):
                sheet.append(row)
            wb.create_sheet("Price_File")
            ws = wb.worksheets[1]
            for row in dataframe_to_rows(Evidence, index=False, header=True):
                ws.append(row)
            wb.save(os.path.join("W:\Admin\Claims Db\Pricing Claims\Tesco\Bacardi",filename))
        elif Retailer_Code=="COO01":
            # i.columns=Calculation_Col
            filename=i["product_number"].iloc[0]+str(i["pricing_date"].iloc[0])[:10]+'.xlsx'
            Prod_No=i["product_number"].iloc[0]
            PF_df_ev=pd.read_sql(query_PF,PF_conn)
            Evidence=PF_df_ev.loc[PF_df_ev["Product_No"]==Prod_No.replace(" ","")]
            wb = Workbook("Calculation")
            wb.create_sheet("Calculation")
            sheet = wb.active
            for row in dataframe_to_rows(i, index=False, header=True):
                sheet.append(row)
            wb.create_sheet("Price_File")
            ws = wb.worksheets[1]
            for row in dataframe_to_rows(Evidence, index=False, header=True):
                ws.append(row)
            wb.save(os.path.join("W:\Admin\Claims Db\Pricing Claims\COOP\Bacardi",filename))

def Evidence_func(PDF):
    Evidence_list=[]
    if os.path.exists(os.path.join(folder.format("NLF"),PDF)):
        images = convert_from_path(os.path.join(folder.format("NLF"),PDF),poppler_path=pop_path)
        for i in range(len(images)):# Save pages as images in the pdf
            images[i].save(folder.format("NLF")+"\\"+PDF.replace(".pdf","_")+ str(i) +'.jpg', 'JPEG')
            Evidence_list.append(folder.format("NLF")+"\\"+PDF.replace(".pdf","_")+ str(i) +'.jpg')
    elif os.path.exists(os.path.join(folder.format("CPC"),PDF)):
        images = convert_from_path(os.path.join(folder.format("CPC"),PDF),poppler_path=pop_path)
        for i in range(len(images)):# Save pages as images in the pdf
            images[i].save(folder.format("CPC")+"\\"+PDF.replace(".pdf","_")+ str(i) +'.jpg', 'JPEG')
            Evidence_list.append(folder.format("CPC")+"\\"+PDF.replace(".pdf","_")+ str(i) +'.jpg')
    return Evidence_list

def insert_image(File,images,df):
    wb = Workbook("Calculation")
    wb.create_sheet("Calculation")
    sheet = wb.active
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)
    for i in range(len(images)):
        wb.create_sheet("Price file "+str(i+1))
        ws = wb.worksheets[i+1]
        img = openpyxl.drawing.image.Image(images[i])
        ws.add_image(img,'A1')
    wb.save(os.path.join("W:\Admin\Claims Db\Pricing Claims\Sainsbury\Bacardi",File))

SI_Analysis_df=Insert_Correct_Case_Cost(Sales_Invoices_df,PF_df)
Claim_Identifier(SI_Analysis_df)
SI_Analysis_df.to_excel('Oli_Test.xlsx',index=False)
